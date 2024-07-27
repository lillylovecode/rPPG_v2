# -*- coding: utf-8 -*-
"""
Created on Mon Mar 21 20:02:23 2022

@author: a1016
"""
import datetime
import os
from openpyxl import Workbook,load_workbook
import time
import numpy as np
import zipfile

# 檔案儲存路徑
basePath = r'C:/rPPG_outcome//'

def zip_dir(path,out_path):
    zf = zipfile.ZipFile('{}.zip'.format(out_path), 'w', zipfile.ZIP_DEFLATED)
   
    for root, dirs, files in os.walk(path):
        for file_name in files:
            zf.write(os.path.join(root, file_name))

# 檔案寫入功能模組
def write_to_file(file_path, data):
    file = open(file_path, "a")  # 使用 'a' 模式為在檔案末端附加新內容
    np.savetxt(file, data.flatten())
    file.close()
            
'''          
def saving(filename,result_data,selfsubject_num,selflf_data,selfhf_data,selflfhf_data,
           selflf_array,selfhf_array,selflfhf_array,selfdlfhf_data,
           selfSBP_DBP_data,selfSBP_DBP2_data,selfSampEn_data,selfApEn_data,
           selfSDNN_data,selfRMSSD_data):
'''

# 將計算結果寫入Excel檔案
def saving(filename, result_data, selfsubject_num, selflf_data, selfhf_data, selflfhf_data,
           selflf_array, selfhf_array, selflfhf_array, selfdlfhf_data,
           selfSBP_DBP_data, selfSBP_DBP2_data, selfSampEn_data, selfApEn_data,
           selfSDNN_data, selfRMSSD_data, selfSpO2_data, selfR_data,
           result2_data, selfSBP_DBP_data_avg, selfSBP_DBP2_data_avg):
    
    loc_dt = datetime.datetime.today() 
    
    if os.path.isfile(basePath + filename + '.xlsx'):
        load_workbook(basePath + filename + '.xlsx')
    else:
        Workbook().save(basePath + filename + '.xlsx')
        #print("建立新檔案!!!")
        wb = load_workbook(basePath + filename + '.xlsx')
        ws = wb.active
        ws.title="基本測試結果"
        
        '''
        result_title=["學號/編號","姓名","性別",
                      "身高","體重","BMI",
                      "坐高","光度",
                      "(實測)心律","(實測)收縮壓","(實測)舒張壓",
                      "(rPPG)心律","(rPPG)收縮壓","(rPPG)舒張壓","眨眼次數",
                      "LF(nu)平均","HF(nu)平均","LF/HF平均","d(lf/hf)/dt","樣本熵","近似熵",
                      "SDNN","RMSSD","Spo2","數據蒐集總時間","測試時刻"]
        '''
        
        result_title=["學號/編號","姓名","性別",
                      "(rPPG)心率","(rPPG)收縮壓","(rPPG)舒張壓","眨眼次數",
                      "LF(nu)平均","HF(nu)平均","LF/HF平均","d(lf/hf)/dt","樣本熵","近似熵",
                      "SDNN","RMSSD","Spo2", "R", "數據蒐集總時間","測試時刻"]
        
        ws.append(result_title)

        wb.create_sheet("LF(nu)")
        wb.create_sheet("HF(nu)")
        wb.create_sheet("LFHF比值")
        wb.create_sheet("LFHF比值變化量")                        
        wb.create_sheet("SampEn")
        wb.create_sheet("ApEn")
        wb.create_sheet("(rPPG)收縮壓")
        wb.create_sheet("(rPPG)舒張壓")
        wb.create_sheet("SDNN")
        wb.create_sheet("RMSSD")
        wb.create_sheet("SpO2")
        wb.create_sheet("R")
        
        result2_title=["學號/編號", "收縮壓平均線性", "舒張壓平均線性"]
        wb.create_sheet("收縮舒張模型").append(result2_title)
        
        wb.save(basePath + filename + '.xlsx')
    
    wb = load_workbook(basePath + filename + '.xlsx')
    ws1 = wb["基本測試結果"]
    ws2 = wb["LF(nu)"]
    ws3 = wb["HF(nu)"]
    ws4 = wb["LFHF比值"]
    ws5 = wb["LFHF比值變化量"]
    ws6 = wb["SampEn"]
    ws7 = wb["ApEn"]
    ws8 = wb["(rPPG)收縮壓"]
    ws9 = wb["(rPPG)舒張壓"]
    ws10 = wb["SDNN"]
    ws11 = wb["RMSSD"]
    ws12 = wb["SpO2"]
    ws13 = wb["R"]
    ws14 = wb["收縮舒張模型"]
   
    #寫入基本測試結果
    
    ws1.append(result_data)
    #print("寫入基本測試結果!!!")
    
    subject_num = [selfsubject_num]
    lf_data = subject_num + selflf_array
    hf_data = subject_num + selfhf_array
    lfhf_data = subject_num + selflfhf_array
    dlfhf_data = subject_num + selfdlfhf_data[1:]
    
    SampEn_data = subject_num + selfSampEn_data
    ApEn_data = subject_num + selfApEn_data
    
    SBP_DBP_data = subject_num + selfSBP_DBP_data
    SBP_DBP2_data = subject_num + selfSBP_DBP2_data
    
    SDNN_data = subject_num + selfSDNN_data
    RMSSD_data = subject_num + selfRMSSD_data
    
    SpO2_data = subject_num + selfSpO2_data
    R_data = subject_num + selfR_data
   
    #平均數
    BP_avg = selfSBP_DBP_data
    BP11 = np.mean(BP_avg)
    #SBP_avg = subject_num + list(BP11)
    #BP2_avg = SBP_DBP2_data
    BP2_avg = selfSBP_DBP2_data
    BP12 = np.mean(BP2_avg)
    #DBP_avg = subject_num + BP12
    
    #中位數
    #BP_med = SBP_DBP_data
    BP_med = selfSBP_DBP_data
    BP21 = np.median(BP_med)
    #SBP_med = subject_num + BP21
    #BP2_med = SBP_DBP2_data
    BP2_med = selfSBP_DBP2_data
    BP22 = np.median(BP2_med)
    #DBP_med = subject_num + BP22
    
    #收縮壓+舒張壓(平均數線性)
    SBP_DBP_data_avg = 0.161*(BP11)+95.997
    SBP_DBP2_data_avg = -0.025*(BP12)+74.754
    
    #收縮壓+舒張壓(中位數線性)
    SBP_DBP_data_med = 0.102*(BP21) + 101.554
    SBP_DBP2_data_med = -0.042*(BP22) + 76.040


    #寫入LF(nu)
    ws2.append(lf_data)
    
    #寫入HF(nu)
    ws3.append(hf_data)
    
    #寫入LF/HF
    ws4.append(lfhf_data)
    
    #寫入d(LF/HF)/dt 多數據
    ws5.append(dlfhf_data)
    
    #寫入樣本熵
    ws6.append(SampEn_data)
    
    #寫入近似熵
    ws7.append(ApEn_data)
    
    #寫入收縮壓
    ws8.append(SBP_DBP_data)
    
    #寫入舒張壓
    ws9.append(SBP_DBP2_data)
    
    #寫入SDNN
    ws10.append(SDNN_data)
    
    #寫入RMSSD
    ws11.append(RMSSD_data)
    
    #寫入SpO2
    ws12.append(SpO2_data)
    
    #寫入SpO2的R
    ws13.append(R_data)
    
    #寫入收縮舒張壓模型
    ws14.append(result2_data)
    
    SBP_DBP_data_avg = SBP_DBP_data_avg
    SBP_DBP2_data_avg = selfSBP_DBP2_data_avg
    
    '''
    ws12['A1'].value = str(subject_num)
    ws12['B1'].value = '輸出值'
    
  
    ws12['A1'].value = '收縮壓平均線性'
    ws12['B1'].value = SBP_DBP_data_avg
    ws12['A2'].value = '舒張壓平均線性'
    ws12['B2'].value = SBP_DBP2_data_avg
    '''

    
    #print("寫入多數據檔案!!!") 
    #儲存檔案
    wb.save(basePath + filename + '.xlsx')

    
    #壓縮檔案 
                    
    #zip_loc_dt_format = loc_dt.strftime("%Y_%m%d_%H%M")
    #zip_date = str(zip_loc_dt_format)
    #zip_path = r'D:\fatigue'
    #zip_out_path = r'D:\fatigue_records\\'+zip_date
    #zip_dir(zip_path,zip_out_path)
    
    
    #上傳檔案
    #if self.subject_upload == True:                        
        #up.upload(filename)
    
    #寫入完成
    return True


# 將回饋數據寫入Excel檔案
def write_feedback_data(filename, feedback_data):
    # 如果檔案不存在，則建立新檔案
    if not os.path.isfile(basePath + filename + '.xlsx'):
        Workbook().save(basePath + filename + '.xlsx')
        # 讀取Excel檔案
        wb = load_workbook(basePath + filename + '.xlsx')
        ws = wb.active
        
        # 如果工作表不存在，則建立新工作表
        ws.title = "回饋數據"
        if not ws.title == "回饋數據":
            wb.create_sheet("回饋數據")
            ws = wb["回饋數據"]
        # 寫入標題
        ws.append(['編號','時間','HR','SBP','DBP','RR','Delirium','Hb','Bilirubin','BloodTime','Spo2','Glucose'])
        # 將時間欄位加寬
        ws.column_dimensions['B'].width = 20

    else:
        # 讀取Excel檔案
        wb = load_workbook(basePath + filename + '.xlsx')
        ws = wb.active
        ws = wb["回饋數據"]

    # 將回饋數據寫入Excel檔案
    ws.append(feedback_data)
    
    # 儲存檔案
    wb.save(basePath + filename + '.xlsx')
    
    # 寫入完成
    return True